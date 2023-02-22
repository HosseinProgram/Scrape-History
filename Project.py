from datetime import datetime , timedelta
import requests
from copy import copy
import json
import os
from persiantools.jdatetime import JalaliDate
from pprintpp import pprint
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.utils.cell import get_column_letter
import sys 

symbol = Symbols=["فاسمین","فخوز","فملی","فولاد","کرماشا","شفن","شاراک","دکیمی","دجابر","ستران","سرود","ساروم","حتاید","بترانس","حفاری","کچاد","کگل","کروی","اخابر","خبهمن","شپنا","شبهرن","شبریز"]
startdate = [1398,1,1]
enddate = [1400,12,29]
From = [9,0,0]
To =[12,30,0]
StepTime=600 #Seconds

thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

def SetFontAndCell(sizef, sizec, From, To):
    fontStyle = Font(size=str(sizef))

    for j in range(From[1], To[1] + 1):
        for i in range(From[0], To[0] + 1):
            sheet.cell(row=i, column=j).font = fontStyle
        sheet.column_dimensions[get_column_letter(j)].width = sizec

def MergeAndBorder(From, To):
    for i in range(From[0], To[0] + 1):
        for j in range(From[1], To[1] + 1):
            sheet.cell(row=i, column=j).border = thin_border
    sheet.merge_cells(start_row=From[0], start_column=From[1], end_row=To[0], end_column=To[1])

def BorderRange(From, To):
    for i in range(From[0], To[0] + 1):
        for j in range(From[1], To[1] + 1):
            sheet.cell(row=i, column=j).border = thin_border

def AllinMent(From, To):
    for i in range(From[0], To[0] + 1):
        for j in range(From[1], To[1] + 1):
            alignment_obj = copy(sheet.cell(row=i, column=j).alignment)
            alignment_obj.horizontal = 'center'
            alignment_obj.vertical = 'center'
            sheet.cell(row=i, column=j).alignment = alignment_obj


headers={"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36 Edg/109.0.1518.61","Accept-Language":"en-US,en;q=0.9,fa;q=0.8","Connection": "keep-alive","Cookie": "ASP.NET_SessionId=kislljlalcplvzmn2q2ycni0"}
module_dir = os.path.dirname(__file__)
file_path = os.path.join(module_dir, 'InsCodeDict.json')

with open(file_path,encoding='utf-8') as json_file:
    NamadDict=json.load(json_file)

for symbol in Symbols :
    print(symbol)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    Start_Date=JalaliDate(startdate[0], startdate[1] , startdate[2]).to_gregorian()
    Start_Date=datetime(Start_Date.year, Start_Date.month, Start_Date.day)
    Start_Date=Start_Date.replace(hour=From[0],minute=From[1],second=From[2])

    End_Date = JalaliDate(enddate[0],enddate[1],enddate[2]).to_gregorian()
    End_Date = datetime(End_Date.year, End_Date.month, End_Date.day)

    index=NamadDict[symbol]
    datestartindex=int(str(Start_Date.year)+("0"+str(Start_Date.month))[-2:]+("0"+str(Start_Date.day))[-2:])
    dateendindex=int(str(End_Date.year)+("0"+str(End_Date.month))[-2:]+("0"+str(End_Date.day))[-2:])

    OrderRanking = [{},{},{},{},{}]
    ActiveList=requests.get("http://www.tsetmc.com/tsev2/data/InstTradeHistory.aspx?i=%s&Top=999999&A=1"%index).text.split(";")
    ActiveDates=list()
    for date in ActiveList:
        try: 
            ActiveDates.append(int(date.split("@")[0]))
        except: 
            pass
    ActiveDates.reverse()
    for i in range(0,len(ActiveDates)):
        date=ActiveDates[i]
        if date >= datestartindex:
            i1=i
            break
    for i in range(0,len(ActiveDates)):
        date=ActiveDates[i]
        if date >= dateendindex:
            i2=i
            break

    AllBourds=dict()

    j=i1
    while j <=i2:
        dateindex = ActiveDates[j]
        dateindex2=str(dateindex)
        print(JalaliDate.to_jalali(int(dateindex2[0:4]) ,int(dateindex2[4:6]) ,int(dateindex2[6:8])).strftime("%Y/%m/%d"))
        try:
            BestLimits = requests.get('http://cdn.tsetmc.com/api/BestLimits/%s/%s'%(index,dateindex),headers=headers).json()["bestLimitsHistory"]
            CurrentClock = copy(Start_Date)
            EndClock = copy(Start_Date.replace(hour=To[0],minute=To[1],second=To[2]))
            Boards = list()
            while CurrentClock<=EndClock :
                CurrentHEven = int(str(CurrentClock.hour) + ("0"+str(CurrentClock.minute))[-2:] +("0"+str(CurrentClock.second))[-2:])

                maxBuy=0
                minSale = 0
                for limit in BestLimits:
                    if CurrentHEven >= limit["hEven"] :
                        OrderRanking[limit["number"]-1]=limit
                    else:
                        break
                Boards.append([CurrentHEven,OrderRanking[0]["pMeDem"],OrderRanking[0]["pMeOf"]])
                CurrentClock = CurrentClock + timedelta(seconds=600)
            AllBourds[str(dateindex)]=Boards
            j+=1
        except Exception as e :
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
            print(e)

    # Boards=AllBourds[list(AllBourds.keys())[0]]
    BorderRange([1,1], [len(list(AllBourds.keys()))+2,len(Boards)*2+1])
    AllinMent([1,1], [len(list(AllBourds.keys()))+2,len(Boards)*2+1])
    sheet.cell(1,1).value="Time"
    sheet.cell(2,1).value="Date"
    for i in range(0,len(Boards)):
        boardstr=str(Boards[i][0])
        clock=boardstr[0:-4]+":"+boardstr[-4:-2]+":"+boardstr[-2:]
        sheet.cell(1,2*(i+1)).value=clock
        MergeAndBorder([1,2*(i+1)], [1,2*(i+1)+1])
        sheet.cell(2,2*(i+1)).value="Buy"
        sheet.cell(2,2*(i+1)+1).value="Sale"


    for j in range(0,len(list(AllBourds.keys()))):
        dateindex = list(AllBourds.keys())[j]
        persiandate=JalaliDate.to_jalali(int(dateindex[0:4]) ,int(dateindex[4:6]) ,int(dateindex[6:8])).strftime("%Y/%m/%d")
        sheet.cell(j+3,1).value=persiandate
        Boards=AllBourds[dateindex]
        for i in range(0,len(Boards)):
            board=Boards[i]
            sheet.cell(j+3,2*(i+1)).value=board[1]
            sheet.cell(j+3,2*(i+1)+1).value=board[2]
    sheet.column_dimensions['A'].width=12
    sheet.freeze_panes ="A3"
    workbook.save(symbol+".xlsx")



